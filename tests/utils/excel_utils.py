import openpyxl
def read_data(file_path, sheet_name, row, column):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    return sheet.cell(row=row, column=column).value

def write_data(file_path, sheet_name, row, column, data):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
    worksheet.cell(row=row, column=column, value=data)
    workbook.save(file_path)